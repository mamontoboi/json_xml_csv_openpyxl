# Створіть таблицю «матеріали» з таких полів: ідентифікатор, вага, висота та додаткові характеристики матеріалу.
# Поле «додаткові  характеристики матеріалу» має зберігати у собі масив, кожен елемент якого є кортежем із двох значень:
# перше – назва характеристики, а друге – її значення.

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from random import randint

wb = Workbook()
headers = ['№', 'Weight', 'Height', 'Specs']

ws = wb.active
ws.title = 'Materials'

ws.append(headers)
for col in range(1, 5):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, italic=True, color="607D3B")

# filling up the first column
for i in range(2, 7):
    ws.append([i - 1])

# # filling up the range with random value
for i in range(2, 7):
    for j in range(2, 4):
        ws.cell(i, j).value = randint(20, 500)

arr = [("colour", "red"), ("price", "high"), ("horse", "neigh"), ("duck", "quack"), ("roses", "are blue")]

starting_row = 2
starting_col = 4
for row, tup in enumerate(arr):
    for col, val in enumerate(tup):
        ws.cell(row=row+starting_row, column=col+starting_col).value = val

# Для таблиці «матеріалу» з завдання 4 створіть користувальницьку агрегатну функцію,
# яка рахує середнє значення ваги всіх матеріалів вислідної вибірки й округляє значення до цілого.

ws["A7"].value = "Average:"
ws["B7"].value = f'=ROUND((SUM(B2:B6))/{len("B2:B6")}, 0)'

# Для таблиці «матеріалу» з завдання 4 створіть функцію користувача, яка приймає необмежену кількість полів
# і повертає їх конкатенацію.


wb.save('Materials.xlsx')


def conc(col_start, row_start, col_end, row_end):
    """
    :param col_start: Write upper left cell column number
    :param row_start: Write upper left cell row number
    :param col_end: Write lower right cell column number
    :param row_end: Write lower right cell row number
    :return: Concatenated string of all cells
    """
    res = ''
    for j in range(row_start, row_end + 1):
        for i in range(col_start, col_end + 1):
            res += str(ws.cell(j, i).value)
    return res


print(conc(2, 2, 4, 6))
