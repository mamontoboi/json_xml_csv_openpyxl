from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from random import randint

wb = Workbook()  # creating of workbook
ws = wb.active  # creation of sheet
ws.title = "Materials"  # name of the sheet
ws.append(['№', 'Weight', 'Height', 'Addit. specs'])  # creating of headers in the row 1
for row in range(6):
    ws.append([1, 60, 20, 'Hell'])  # to write in the next available row
ws.merge_cells("A2:D2")  # to merge cells
ws.unmerge_cells("A2:D2")  # to unmerge cells
ws.insert_rows(5)  # to insert row in row # 5
ws.insert_cols(3)  # to insert column in column C

ws.delete_rows(5)  # to delete row # 5
ws.delete_cols(3)  # to delete column C

ws.move_range("A3:D4", cols=5, rows=3)  # to move range
ws.move_range("F6:I7", cols=-5, rows=-3)

char = get_column_letter(3)  # to obtain letter on the column knowing it's number
print(char)  # C

# to add function into the cell
ws["B8"].value = "Average:"
ws["C8"].value = f'=SUM(C2:C7)/{len("C2:C7")}'

# to change the style of cell. Only for cell, not for range
ws["B8"].font = Font(bold=True, italic=True, color="FFA07A")

for i in range(2, 12):
    for j in range(2, 5):
        ws.cell(i, j).value = randint(20, 500)


wb.save('exmpl.xlsx')
